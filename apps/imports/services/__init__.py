import os
import re
import zipfile
import tempfile
import shutil
import datetime
import openpyxl
import logging
from io import BytesIO
import mimetypes
from django.conf import settings
from django.utils import timezone
from django.db import transaction, IntegrityError
from rest_framework.exceptions import ValidationError
import hashlib

from apps.fields.models import FieldType, Field
from apps.tables.models import Table
from apps.cards.models import Card, CardUniqueValue
from apps.cards.services import CardService
from apps.imports.models import (
    ImportSession, ImportRowResult, ImportWarning, ReuploadSession, ImportStatus, RowStatus
)
from apps.mediafiles.services import MediaService
from apps.mediafiles.models import MediaFile
from apps.imports.services.metadata import ImageMetadataService
from apps.auditlogs.models import AuditLog

logger = logging.getLogger(__name__)

class ImportService:
    @staticmethod
    def detect_columns_from_excel(file_path: str) -> list:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if not wb.sheetnames:
            raise ValidationError("Excel file has no sheets.")
        ws = wb.active
        
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            raise ValidationError("Excel file is empty.")
            
        if not headers:
            raise ValidationError("Excel file has no headers.")
            
        headers = [str(h).strip() for h in headers if h is not None]
        
        sample_data = []
        for _ in range(100):
            try:
                row = next(rows_iter)
                sample_data.append(row)
            except StopIteration:
                break
                
        preview = []
        for col_idx, header in enumerate(headers):
            col_values = []
            for row in sample_data:
                if col_idx < len(row):
                    val = row[col_idx]
                    if val is not None:
                        col_values.append(val)
                        
            detected_type = FieldType.TEXT
            confidence = 1.0
            is_required = len(col_values) == len(sample_data) and len(sample_data) > 0
            
            unique_vals = set(col_values)
            is_unique = len(unique_vals) == len(col_values) and len(col_values) > 0
            
            if col_values:
                type_counts = {
                    FieldType.TEXT: 0,
                    FieldType.NUMBER: 0,
                    FieldType.DATE: 0,
                    FieldType.BOOLEAN: 0,
                    FieldType.IMAGE: 0,
                }
                
                for val in col_values:
                    val_str = str(val).strip()
                    if any(val_str.lower().endswith(ext) for ext in ['.png', '.jpg', '.jpeg', '.webp']):
                        type_counts[FieldType.IMAGE] += 1
                    elif isinstance(val, (datetime.date, datetime.datetime)):
                        type_counts[FieldType.DATE] += 1
                    elif re.match(r'^\d{4}-\d{2}-\d{2}(\s+\d{2}:\d{2}:\d{2})?$', val_str):
                        type_counts[FieldType.DATE] += 1
                    elif isinstance(val, (int, float)) and not isinstance(val, bool):
                        type_counts[FieldType.NUMBER] += 1
                    elif re.match(r'^-?\d+(\.\d+)?$', val_str):
                        type_counts[FieldType.NUMBER] += 1
                    elif isinstance(val, bool):
                        type_counts[FieldType.BOOLEAN] += 1
                    elif val_str.lower() in ['true', 'false', 'yes', 'no', '1', '0']:
                        type_counts[FieldType.BOOLEAN] += 1
                    else:
                        type_counts[FieldType.TEXT] += 1
                
                best_type = FieldType.TEXT
                best_count = type_counts[FieldType.TEXT]
                
                for t, count in type_counts.items():
                    if t != FieldType.TEXT and count > best_count:
                        best_count = count
                        best_type = t
                        
                detected_type = best_type
                confidence = best_count / len(col_values)
                
            preview.append({
                "column_name": header,
                "detected_type": detected_type,
                "confidence": round(confidence, 2),
                "is_required": is_required,
                "is_unique": is_unique
            })
            
        return preview

    @staticmethod
    def process_xlsx_import(
        import_session: ImportSession,
        excel_path: str,
        zip_path: str = None,
        job=None,
        progress_callback=None
    ):
        """
        Processes an XLSX import inside the Job / Celery infrastructure.
        Ensures bad rows do not fail the import, and duplicates produce ImportWarnings.
        """
        import_session.status = ImportStatus.PROCESSING
        import_session.started_at = timezone.now()
        import_session.save()
        
        # Log Audit event
        AuditLog.objects.create(
            event_type='IMPORT_START',
            actor=import_session.user,
            target_organization=import_session.organization,
            details={"import_session_id": str(import_session.id), "table_id": str(import_session.table_id)}
        )

        # 1. Unzip images if ZIP path is provided
        extracted_zip_dir = None
        if zip_path and os.path.exists(zip_path):
            if progress_callback:
                progress_callback(10, "Extracting zip archive")
            extracted_zip_dir = tempfile.mkdtemp()
            try:
                with zipfile.ZipFile(zip_path, 'r') as zf:
                    zf.extractall(extracted_zip_dir)
            except Exception as e:
                logger.warning(f"Failed to extract zip file: {e}")
                
        # 2. Open workbook in read-only mode to avoid memory spikes
        if progress_callback:
            progress_callback(20, "Reading Excel workbook")
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
        except StopIteration:
            import_session.status = ImportStatus.FAILED
            import_session.completed_at = timezone.now()
            import_session.save()
            return
            
        headers = [str(h).strip() for h in headers if h is not None]
        fields = Field.objects.filter(table=import_session.table, is_deleted=False)
        field_by_name = {f.name.lower(): f for f in fields}
        
        # 3. Read and group rows in chunks of 500
        chunk_size = 500
        current_chunk = []
        row_num = 1  # 1 is header row
        
        total_rows = 0
        success_rows = 0
        warning_rows = 0
        failed_rows = 0
        
        for row in rows_iter:
            row_num += 1
            if not any(val is not None for val in row):
                continue  # skip completely blank rows
            total_rows += 1
            current_chunk.append((row_num, row))
            
            if len(current_chunk) >= chunk_size:
                s, w, f = ImportService._import_chunk_transaction(
                    import_session, current_chunk, headers, field_by_name, extracted_zip_dir
                )
                success_rows += s
                warning_rows += w
                failed_rows += f
                current_chunk = []
                
                # Check cancellation midway
                if job:
                    job.refresh_from_db()
                    if job.status == 'CANCELLED':
                        import_session.status = ImportStatus.FAILED
                        import_session.completed_at = timezone.now()
                        import_session.save()
                        # Clean up temp zip extraction
                        if extracted_zip_dir:
                            shutil.rmtree(extracted_zip_dir, ignore_errors=True)
                        return
                
                if progress_callback:
                    pct = min(20 + int((row_num / 5000) * 60), 80)
                    progress_callback(pct, f"Processed {row_num} rows")

        # Process any remaining rows in the last chunk
        if current_chunk:
            s, w, f = ImportService._import_chunk_transaction(
                import_session, current_chunk, headers, field_by_name, extracted_zip_dir
            )
            success_rows += s
            warning_rows += w
            failed_rows += f

        # Clean up zip extraction directory
        if extracted_zip_dir:
            shutil.rmtree(extracted_zip_dir, ignore_errors=True)
            
        # 4. Finalize session stats
        completed_time = timezone.now()
        import_session.status = ImportStatus.COMPLETED
        import_session.completed_at = completed_time
        import_session.total_rows = total_rows
        import_session.success_rows = success_rows
        import_session.warning_rows = warning_rows
        import_session.failed_rows = failed_rows
        
        if import_session.started_at:
            import_session.duration = (completed_time - import_session.started_at).total_seconds()
            
        import_session.save()
        
        # Log Audit event
        AuditLog.objects.create(
            event_type='IMPORT_COMPLETE',
            actor=import_session.user,
            target_organization=import_session.organization,
            details={
                "import_session_id": str(import_session.id),
                "total_rows": total_rows,
                "success_rows": success_rows,
                "warning_rows": warning_rows,
                "failed_rows": failed_rows
            }
        )

    @staticmethod
    @transaction.atomic
    def _import_chunk_transaction(import_session, chunk, headers, field_by_name, extracted_zip_dir) -> tuple:
        """
        Imports a chunk of rows inside a database transaction block.
        If a row fails completely, we record the failure, but we do not roll back the entire chunk.
        """
        success_count = 0
        warning_count = 0
        failed_count = 0

        table = import_session.table
        org_id = str(import_session.organization_id)

        # Pre-build field_by_id map to avoid N+1 queries inside the loop
        all_fields = Field.objects.filter(table=table, is_deleted=False)
        field_by_id = {str(f.id): f for f in all_fields}

        # Lock table row once for this entire chunk to get sequential display_ids atomically
        from django.db.models import F
        from apps.tables.models import Table as TableModel
        from apps.cards.models import CardStatus

        locked_table = TableModel.objects.select_for_update().get(pk=table.pk)

        for row_num, row_values in chunk:
            row_data = {}
            # Build data mapped to fields
            for col_idx, val in enumerate(row_values):
                if col_idx < len(headers):
                    header = headers[col_idx].lower()
                    if header in field_by_name:
                        field = field_by_name[header]
                        row_data[str(field.id)] = val

            # Start row processing
            row_warnings = []
            row_failed = False
            error_message = None

            # 1. Image field pre-processing (placeholder until card is created)
            processed_row_data = {}
            for field_id, val in row_data.items():
                field = field_by_id.get(field_id)
                if field and field.type == FieldType.IMAGE:
                    val_str = str(val or '').strip()
                    if val_str:
                        # Optimistic: mark as PENDING_IMAGE until we associate post-card-creation
                        processed_row_data[field_id] = "PENDING_IMAGE_PLACEHOLDER"
                        if not extracted_zip_dir:
                            row_warnings.append(("MISSING_IMAGE", f"Image path '{val_str}' specified but no ZIP provided."))
                    else:
                        processed_row_data[field_id] = "NO_IMAGE_PLACEHOLDER"
                else:
                    processed_row_data[field_id] = val

            # 2. Pre-validate Uniqueness (duplicates allowed with warnings)
            unique_violations = []
            for field_id, val in list(processed_row_data.items()):
                field = field_by_id.get(field_id)
                if field and field.is_unique and val is not None and str(val).strip() not in ("", "PENDING_IMAGE_PLACEHOLDER", "NO_IMAGE_PLACEHOLDER"):
                    value_hash = hashlib.sha256(str(val).encode('utf-8')).hexdigest()
                    if CardUniqueValue.objects.filter(table_id=table.id, field_id=field.id, value_hash=value_hash).exists():
                        unique_violations.append(field)

            # Log warnings for duplicates
            for uf in unique_violations:
                row_warnings.append(("DUPLICATE", f"Duplicate value for unique field '{uf.name}' detected."))

            # 3. Create Card with atomic display_id
            try:
                locked_table.card_sequence = F('card_sequence') + 1
                locked_table.save(update_fields=['card_sequence'])
                locked_table.refresh_from_db(fields=['card_sequence'])
                display_id = f"TBL-{locked_table.card_sequence}"

                card = Card.objects.create(
                    table=table,
                    organization_id=org_id,
                    display_id=display_id,
                    status=CardStatus.PENDING,
                    data=processed_row_data,
                    created_by=import_session.user
                )

                # Write unique values for non-violating fields
                for field_id, val in processed_row_data.items():
                    field = field_by_id.get(field_id)
                    if field and field.is_unique and field not in unique_violations and val is not None and str(val).strip() not in ("", "PENDING_IMAGE_PLACEHOLDER", "NO_IMAGE_PLACEHOLDER"):
                        value_hash = hashlib.sha256(str(val).encode('utf-8')).hexdigest()
                        CardUniqueValue.objects.create(
                            card=card,
                            table_id=table.id,
                            field_id=field.id,
                            value_hash=value_hash
                        )

                # 4. Associate ZIP images now that card is created
                for field_id, val in row_data.items():
                    field = field_by_id.get(field_id)
                    if field and field.type == FieldType.IMAGE and str(val or '').strip() and extracted_zip_dir:
                        val_str = str(val).strip()
                        matched_file_path = None
                        possible_paths = [
                            os.path.join(extracted_zip_dir, val_str),
                            os.path.join(extracted_zip_dir, os.path.basename(val_str)),
                        ]
                        for p in possible_paths:
                            if os.path.exists(p) and os.path.isfile(p):
                                matched_file_path = p
                                break
                        if matched_file_path:
                            try:
                                with open(matched_file_path, 'rb') as f:
                                    mime_type, _ = mimetypes.guess_type(matched_file_path)
                                    MediaService.upload_image(
                                        card=card,
                                        field=field,
                                        file_name=os.path.basename(val_str),
                                        content_type=mime_type or "image/jpeg",
                                        file_content=f,
                                        user=import_session.user
                                    )
                            except Exception as e:
                                row_warnings.append(("IMAGE_UPLOAD_FAIL", f"Failed to upload matched image '{val_str}': {e}"))
                        else:
                            row_warnings.append(("MISSING_IMAGE", f"Image '{val_str}' not found in ZIP archive."))

            except Exception as e:
                row_failed = True
                error_message = str(e)

            # 5. Log Row results and warnings
            if row_failed:
                failed_count += 1
                ImportRowResult.objects.create(
                    import_session=import_session,
                    row_number=row_num,
                    status=RowStatus.FAILED,
                    error_message=error_message,
                    row_data=row_data
                )
            else:
                success_count += 1
                ImportRowResult.objects.create(
                    import_session=import_session,
                    row_number=row_num,
                    status=RowStatus.WARNING if row_warnings else RowStatus.SUCCESS,
                    row_data=row_data
                )
                if row_warnings:
                    warning_count += 1
                    for w_type, w_msg in row_warnings:
                        ImportWarning.objects.create(
                            import_session=import_session,
                            row_number=row_num,
                            warning_type=w_type,
                            message=w_msg
                        )

        return success_count, warning_count, failed_count


    @staticmethod
    def process_image_reupload(
        reupload_session: ReuploadSession,
        zip_path: str,
        progress_callback=None
    ):
        """
        Processes a ZIP of reuploaded images.
        Extracts EXIF/PNG metadata to match card IDs, falling back to filename matching.
        """
        reupload_session.status = ImportStatus.PROCESSING
        reupload_session.started_at = timezone.now()
        reupload_session.save()
        
        # Log Audit event
        AuditLog.objects.create(
            event_type='REUPLOAD_START',
            actor=reupload_session.user,
            target_organization=reupload_session.organization,
            details={"reupload_session_id": str(reupload_session.id), "table_id": str(reupload_session.table_id)}
        )

        extracted_dir = tempfile.mkdtemp()
        total_images = 0
        matched_images = 0
        failed_images = 0
        
        try:
            if progress_callback:
                progress_callback(10, "Extracting zip archive")
            with zipfile.ZipFile(zip_path, 'r') as zf:
                zf.extractall(extracted_dir)
                
            # Iterate through all extracted files
            all_files = []
            for root, dirs, files in os.walk(extracted_dir):
                for f in files:
                    if any(f.lower().endswith(ext) for ext in ['.png', '.jpg', '.jpeg', '.webp']):
                        all_files.append(os.path.join(root, f))
                        
            total_images = len(all_files)
            if progress_callback:
                progress_callback(30, f"Found {total_images} images to process")
                
            for idx, file_path in enumerate(all_files):
                filename = os.path.basename(file_path)
                
                # 1. Match using metadata
                matched = False
                try:
                    with open(file_path, 'rb') as f:
                        img_bytes = f.read()
                        meta = ImageMetadataService.extract_metadata(img_bytes)
                        
                    card_id = meta.get("card_id")
                    media_id = meta.get("media_id")
                    
                    if card_id:
                        card = Card.objects.filter(id=card_id, table=reupload_session.table).first()
                        if card:
                            # Determine which image field it was
                            # If media_id is provided, find the field from the MediaFile
                            field = None
                            if media_id:
                                mf = MediaFile.objects.filter(id=media_id).first()
                                if mf:
                                    field = mf.field
                            if not field:
                                # Fallback: find the first Image field on the table
                                field = Field.objects.filter(table=reupload_session.table, type=FieldType.IMAGE).first()
                                
                            if field:
                                with open(file_path, 'rb') as f:
                                    mime_type, _ = mimetypes.guess_type(file_path)
                                    MediaService.replace_image(
                                        card=card,
                                        field=field,
                                        file_name=filename,
                                        content_type=mime_type or "image/png",
                                        file_content=f,
                                        user=reupload_session.user
                                    )
                                matched = True
                                matched_images += 1
                except Exception as e:
                    logger.warning(f"Metadata matching failed for {filename}: {e}")
                    
                # 2. Match using filename (Fallback)
                if not matched:
                    try:
                        # Find an active MediaFile by original name
                        mf = MediaFile.objects.filter(
                            original_name=filename,
                            table=reupload_session.table,
                            is_deleted=False
                        ).first()
                        
                        if mf and mf.card:
                            with open(file_path, 'rb') as f:
                                mime_type, _ = mimetypes.guess_type(file_path)
                                MediaService.replace_image(
                                    card=mf.card,
                                    field=mf.field,
                                    file_name=filename,
                                    content_type=mime_type or "image/png",
                                    file_content=f,
                                    user=reupload_session.user
                                )
                            matched = True
                            matched_images += 1
                    except Exception as e:
                        logger.warning(f"Filename matching failed for {filename}: {e}")
                        
                if not matched:
                    failed_images += 1
                    
                if progress_callback:
                    pct = min(30 + int((idx / max(total_images, 1)) * 60), 90)
                    progress_callback(pct, f"Processed {idx + 1} of {total_images} images")
                    
        finally:
            shutil.rmtree(extracted_dir, ignore_errors=True)
            
        completed_time = timezone.now()
        reupload_session.status = ImportStatus.COMPLETED
        reupload_session.completed_at = completed_time
        reupload_session.total_images = total_images
        reupload_session.matched_images = matched_images
        reupload_session.failed_images = failed_images
        
        if reupload_session.started_at:
            reupload_session.duration = (completed_time - reupload_session.started_at).total_seconds()
            
        reupload_session.save()
        
        # Log Audit event
        AuditLog.objects.create(
            event_type='REUPLOAD_COMPLETE',
            actor=reupload_session.user,
            target_organization=reupload_session.organization,
            details={
                "reupload_session_id": str(reupload_session.id),
                "total_images": total_images,
                "matched_images": matched_images,
                "failed_images": failed_images
            }
        )
