"""
Excel Export Module

Handles XLSX file generation for ID card data.
This module is READ-ONLY - it never mutates data.
"""
import logging
from io import BytesIO
from typing import Dict, Any, Optional
from dataclasses import dataclass

from django.http import HttpResponse
from django.db.models import QuerySet

from .utils import (
    get_text_fields,
    get_image_fields,
    generate_export_filename,
    format_field_value,
    sort_cards_for_export,
    stream_file_response,
)

logger = logging.getLogger(__name__)


@dataclass
class ExcelExportResult:
    """Result of an Excel export operation."""
    success: bool
    message: str = ''
    response: Optional[HttpResponse] = None
    filename: str = ''
    row_count: int = 0


class ExcelExporter:
    """
    Handles Excel (XLSX) export operations.
    
    Features:
    - Exports only text fields (no images)
    - Auto-sizes columns
    - Applies consistent formatting
    - Freezes header row
    
    Usage:
        exporter = ExcelExporter()
        result = exporter.export_cards(table, cards)
        if result.success:
            return result.response
    """
    
    # Excel sheet name max length
    MAX_SHEET_NAME_LENGTH = 31
    
    # Column width limits
    MIN_COLUMN_WIDTH = 8
    MAX_COLUMN_WIDTH = 50
    
    def export_cards(
        self,
        table,
        cards: QuerySet,
        uppercase_values: bool = True,
        status: str = '',
        progress_callback=None,
        user=None,
    ) -> ExcelExportResult:
        """
        Export cards to Excel format.
        
        Args:
            table: IDCardTable instance
            cards: QuerySet of IDCard instances
            uppercase_values: Convert values to uppercase
            
        Returns:
            ExcelExportResult with HttpResponse if successful
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return ExcelExportResult(
                success=False,
                message='openpyxl library not installed. Run: pip install openpyxl'
            )
        
        if not cards.exists():
            return ExcelExportResult(
                success=False,
                message='No cards to export!'
            )
        
        try:
            # Get text fields only (exclude images)
            text_fields = get_text_fields(table.fields or [])
            # Get image fields (for filename-only columns)
            image_fields = get_image_fields(table.fields or [])
            
            if not text_fields and not image_fields:
                return ExcelExportResult(
                    success=False,
                    message='No fields found in table configuration!'
                )
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = table.name[:self.MAX_SHEET_NAME_LENGTH]
            
            # Define styles
            styles = self._get_styles(Font, Alignment, Border, Side)
            
            # Track column widths
            column_widths = {}
            
            # Write header row — text fields first, then image fields
            headers = [f['name'] for f in text_fields] + [f['name'] for f in image_fields]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = styles['header_font']
                cell.alignment = styles['header_alignment']
                cell.border = styles['border']
                column_widths[col_idx] = len(str(header)) + 2
            
            # Sort cards for export (Class → Section → Name)
            sorted_cards = sort_cards_for_export(cards, table.fields)

            # Write data rows
            row_count = 0
            text_col_count = len(text_fields)
            for row_idx, card in enumerate(sorted_cards, 2):
                field_data = card.field_data or {}
                
                # Text field columns
                for col_idx, field in enumerate(text_fields, 1):
                    value = field_data.get(field['name'], '')
                    formatted_value = format_field_value(value, uppercase=uppercase_values)
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=formatted_value)
                    cell.font = styles['data_font']
                    cell.alignment = styles['data_alignment']
                    cell.border = styles['border']
                    
                    # Track max width
                    current_width = min(len(formatted_value) + 2, self.MAX_COLUMN_WIDTH)
                    column_widths[col_idx] = max(
                        column_widths.get(col_idx, self.MIN_COLUMN_WIDTH),
                        current_width
                    )
                
                # Image field columns — show filename without extension
                for img_idx, img_field in enumerate(image_fields):
                    col_idx = text_col_count + img_idx + 1
                    raw_value = field_data.get(img_field['name'], '')
                    filename_stem = self._extract_image_filename(raw_value)
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=filename_stem)
                    cell.font = styles['data_font']
                    cell.alignment = styles['data_alignment']
                    cell.border = styles['border']
                    
                    current_width = min(len(filename_stem) + 2, self.MAX_COLUMN_WIDTH)
                    column_widths[col_idx] = max(
                        column_widths.get(col_idx, self.MIN_COLUMN_WIDTH),
                        current_width
                    )
                
                row_count += 1

                if callable(progress_callback) and ((row_count % 25 == 0) or (row_count == len(sorted_cards))):
                    try:
                        progress_callback(row_count, len(sorted_cards))
                    except Exception:
                        pass
            
            # Apply column widths
            for col_idx, width in column_widths.items():
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(self.MIN_COLUMN_WIDTH, width * 1.1)
            
            # Set header row height and freeze
            ws.row_dimensions[1].height = 25
            ws.freeze_panes = 'A2'
            
            # Save to buffer
            xlsx_buffer = BytesIO()
            wb.save(xlsx_buffer)
            xlsx_buffer.seek(0)
            
            # Get client name for filename
            client_name = ''
            if table.group and table.group.client:
                client_name = table.group.client.name

            # Generate filename
            filename = generate_export_filename(table.name, 'xlsx', client_name=client_name, status=status)
            
            # Create response — uses chunked streaming for large files
            xlsx_bytes = xlsx_buffer.getvalue()
            xlsx_buffer.close()
            response = stream_file_response(
                xlsx_bytes,
                filename,
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                user=user,
            )
            
            return ExcelExportResult(
                success=True,
                response=response,
                filename=filename,
                row_count=row_count
            )
            
        except Exception as e:
            logger.error("Excel export failed: %s", e, exc_info=True)
            return ExcelExportResult(
                success=False,
                message='Excel export failed. Please try again or contact support.'
            )
    
    def _get_styles(self, Font, Alignment, Border, Side) -> Dict[str, Any]:
        """Get style definitions for Excel export."""
        return {
            'header_font': Font(name='Arial', size=11, bold=True),
            'header_alignment': Alignment(horizontal='center', vertical='center'),
            'data_font': Font(name='Arial', size=10),
            'data_alignment': Alignment(horizontal='left', vertical='center', wrap_text=False),
            'border': Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
        }

    @staticmethod
    def _extract_image_filename(raw_value: str) -> str:
        """Extract image filename without extension from a field_data value.

        Handles these formats:
          • 'clients_imgs/Client/table/1234.jpg'  → '1234'
          • 'PENDING:filename.jpg'                → 'filename'
          • 'NOT_FOUND'                           → ''
          • ''                                    → ''
        """
        import os as _os

        if not raw_value or raw_value in ('NOT_FOUND', ''):
            return ''
        val = str(raw_value).strip()
        # Strip PENDING: prefix
        if val.upper().startswith('PENDING:'):
            val = val[8:]
        # Get basename, then strip extension
        basename = _os.path.basename(val)
        name, _ = _os.path.splitext(basename)
        return name

