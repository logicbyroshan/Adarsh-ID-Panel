"""
Backup Service  (panel app)
===========================

Background service that creates a SINGLE combined ZIP archive named
``Adarsh Backup {YYYY-MM-DD}.zip`` containing:
  • One folder per school/client
  • Inside each school folder: XLSX files split by status
    (pending.xlsx, verified.xlsx …) plus an ``images/`` folder
    holding all photos referenced in the cards

Uses a daemon thread for the heavy I/O so the request returns immediately.
Progress is tracked in the BackupTask model (in core.models), polled from
the Manage Panel.
"""

import logging
import os
import threading
import zipfile
from io import BytesIO
from typing import List, Optional

from django.conf import settings
from django.utils import timezone

logger = logging.getLogger(__name__)

# Where backup ZIPs live (relative folder under MEDIA_ROOT)
BACKUP_DIR_NAME = os.path.join('temp', 'backups')


def _backup_root():
    """Absolute path to the backup directory."""
    return os.path.join(settings.MEDIA_ROOT, BACKUP_DIR_NAME)


def _ensure_backup_dir(task_id: int):
    """Create per-task backup directory."""
    d = os.path.join(_backup_root(), str(task_id))
    os.makedirs(d, exist_ok=True)
    return d


# ─── Public API ───────────────────────────────────────────────────────────

def start_backup(task_id: int):
    """
    Launch the background thread that processes the backup.

    ``task_id`` must reference an existing ``BackupTask`` in *pending* state.
    """
    thread = threading.Thread(
        target=_process_backup,
        args=(task_id,),
        daemon=True,
    )
    thread.start()


def delete_backup_files(task_id: int):
    """Immediately delete files for a backup task and mark it deleted."""
    from core.models import BackupTask

    try:
        task = BackupTask.objects.get(pk=task_id)
        task.cleanup_files()
        task.status = 'deleted'
        task.save(update_fields=['status'])
        logger.info("Backup #%d files deleted.", task_id)
    except BackupTask.DoesNotExist:
        logger.warning("delete_backup_files: BackupTask #%d not found", task_id)


# ─── Background worker ───────────────────────────────────────────────────

def _process_backup(task_id: int):
    """Main background thread — iterates clients and builds ONE combined ZIP.

    Creates a single ``Adarsh Backup {YYYY-MM-DD}.zip`` with one folder per
    school/client.
    """
    from core.models import BackupTask
    from client.models import Client

    try:
        task = BackupTask.objects.get(pk=task_id)
    except BackupTask.DoesNotExist:
        logger.error("BackupTask #%d vanished before processing", task_id)
        return

    task.status = 'processing'
    task.started_at = timezone.now()
    task.save(update_fields=['status', 'started_at'])

    out_dir = _ensure_backup_dir(task_id)
    client_ids: List[int] = task.client_ids or []

    try:
        clients = (
            Client.objects
            .filter(pk__in=client_ids)
            .prefetch_related('id_card_groups__tables')
            .order_by('name')
        )
        task.total = clients.count()
        task.save(update_fields=['total'])

        # ── Build ONE combined ZIP ──────────────────────────────────────
        date_str = timezone.now().strftime('%Y-%m-%d')
        zip_filename = f"Adarsh Backup {date_str}.zip"
        zip_path = os.path.join(out_dir, zip_filename)

        has_any_data = False
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as combined_zf:
            for idx, client in enumerate(clients, 1):
                task.current_client = client.name
                task.save(update_fields=['current_client'])

                wrote = _build_client_in_zip(combined_zf, client)
                if wrote:
                    has_any_data = True

                task.progress = idx
                task.save(update_fields=['progress'])

        if has_any_data and os.path.exists(zip_path):
            file_size = os.path.getsize(zip_path)
            rel_path = os.path.relpath(zip_path, settings.MEDIA_ROOT).replace('\\', '/')
            combined_info = {
                'path': rel_path,
                'filename': zip_filename,
                'size': file_size,
            }
        else:
            if os.path.exists(zip_path):
                os.remove(zip_path)
            combined_info = None

        task.zip_files = {'combined': combined_info} if combined_info else {}
        task.progress = task.total
        task.current_client = ''
        task.status = 'completed'
        task.completed_at = timezone.now()
        task.save(update_fields=[
            'zip_files', 'progress', 'current_client',
            'status', 'completed_at',
        ])
        logger.info("Backup #%d completed — combined ZIP: %s", task_id, zip_filename)

    except Exception as exc:
        logger.exception("Backup #%d failed: %s", task_id, exc)
        try:
            task.refresh_from_db()
            task.status = 'failed'
            task.error_message = str(exc)[:2000]
            task.completed_at = timezone.now()
            task.save(update_fields=['status', 'error_message', 'completed_at'])
        except Exception as persist_err:
            logger.warning("Backup #%d failed and status persistence also failed: %s", task_id, persist_err)




def _build_client_in_zip(zf: zipfile.ZipFile, client) -> bool:
    """
    Write one client's data into an already-open combined ZipFile.

    Folder structure::

        {ClientName}/
          {TableName}/
            pending.xlsx
            verified.xlsx
            images/
              photo1.jpg …

    Returns True if at least one file was written.
    """
    from idcards.models import IDCardTable

    groups = list(client.id_card_groups.all())
    if not groups:
        return False

    safe_name = _safe_filename(client.name)
    wrote_any = False

    for group in groups:
        tables = getattr(group, 'tables', IDCardTable.objects.filter(group=group)).all()
        for table in tables:
            wrote = _write_table_to_zip(zf, safe_name, table)
            if wrote:
                wrote_any = True

    return wrote_any


def _write_table_to_zip(zf: zipfile.ZipFile, client_folder: str, table) -> bool:
    """Write one table's data into the ZIP (XLSX per status + images)."""
    from idcards.models import IDCard
    from exports.utils import get_text_fields, get_image_fields

    all_cards = list(IDCard.objects.filter(table=table).order_by('id'))
    if not all_cards:
        return False

    table_safe = _safe_filename(table.name)
    base_prefix = f"{client_folder}/{table_safe}"

    text_fields = get_text_fields(table.fields or [])
    image_fields = get_image_fields(table.fields or [])

    statuses = ['pending', 'verified', 'pool', 'approved', 'download', 'reprint']
    wrote_any = False
    image_paths_written = set()

    cards_by_status = {status: [] for status in statuses}
    for card in all_cards:
        if card.status in cards_by_status:
            cards_by_status[card.status].append(card)

    for status in statuses:
        cards = cards_by_status[status]
        if not cards:
            continue

        xlsx_bytes = _build_xlsx_for_cards(cards, text_fields, image_fields, table.name)
        if xlsx_bytes:
            zf.writestr(f"{base_prefix}/{status}.xlsx", xlsx_bytes)
            wrote_any = True

        for card in cards:
            _collect_images(zf, base_prefix, card, image_fields, image_paths_written)

    return wrote_any


def _build_xlsx_for_cards(cards, text_fields, image_fields, sheet_name: str):
    """Build an in-memory XLSX for a list of cards (single status group)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from exports.utils import format_field_value
    except ImportError:
        logger.error("openpyxl not installed — cannot build XLSX for backup")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    hdr_font = Font(name='Arial', size=11, bold=True)
    hdr_align = Alignment(horizontal='center', vertical='center')
    data_font = Font(name='Arial', size=10)
    data_align = Alignment(horizontal='left', vertical='center')
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    headers = [f['name'] for f in text_fields] + [f['name'] for f in image_fields]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = border

    text_count = len(text_fields)
    for ri, card in enumerate(cards, 2):
        fd = card.field_data or {}
        for ci, f in enumerate(text_fields, 1):
            val = format_field_value(fd.get(f['name'], ''), uppercase=True)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border
        for ii, img_f in enumerate(image_fields):
            ci = text_count + ii + 1
            raw = fd.get(img_f['name'], '')
            fname = _extract_image_stem(raw)
            cell = ws.cell(row=ri, column=ci, value=fname)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border

    for ci in range(1, len(headers) + 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            (len(str(ws.cell(row=r, column=ci).value or '')) for r in range(1, len(cards) + 2)),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(50, max(8, max_len + 2))
    ws.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _collect_images(zf, base_prefix, card, image_fields, already):
    """Add image files referenced by *card* into ``{base_prefix}/images/``."""
    fd = card.field_data or {}
    for f in image_fields:
        raw = fd.get(f['name'], '')
        if not raw or raw in ('NOT_FOUND', ''):
            continue
        if raw.startswith('PENDING:'):
            continue

        rel = raw.replace('\\', '/')
        abs_path = _resolve_media_path(rel)
        if not abs_path or not os.path.isfile(abs_path):
            continue

        arc_name = f"{base_prefix}/images/{os.path.basename(abs_path)}"
        if arc_name in already:
            continue
        already.add(arc_name)

        try:
            zf.write(abs_path, arc_name)
        except Exception as exc:
            logger.debug("Could not add image %s to ZIP: %s", abs_path, exc)


# ─── Helpers ─────────────────────────────────────────────────────────────

def _safe_filename(name: str) -> str:
    """Convert a name into a filesystem-safe string."""
    import re
    safe = re.sub(r'[^\w\s-]', '', name).strip()
    safe = re.sub(r'[\s]+', '_', safe)
    return safe[:80] or 'unnamed'


def _extract_image_stem(raw: str) -> str:
    """Extract filename without extension from image path."""
    if not raw or raw in ('NOT_FOUND', ''):
        return ''
    val = str(raw).strip()
    if val.upper().startswith('PENDING:'):
        val = val[8:]
    basename = os.path.basename(val)
    name, _ = os.path.splitext(basename)
    return name


def _resolve_media_path(rel_path: str) -> Optional[str]:
    """Resolve a media-relative path and reject traversal/absolute escapes."""
    if not rel_path:
        return None

    media_root = os.path.abspath(settings.MEDIA_ROOT)
    candidate = os.path.abspath(os.path.join(media_root, rel_path))
    try:
        if os.path.commonpath([media_root, candidate]) != media_root:
            return None
    except ValueError:
        return None
    return candidate
