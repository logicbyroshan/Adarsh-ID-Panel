"""
ID Card Table API — table CRUD + XLSX import.

Contains:
- api_idcard_table_create, api_idcard_table_get, api_idcard_table_update,
  api_idcard_table_delete, api_idcard_table_toggle_status, api_idcard_table_list
- _HEADER_TYPE_MAP, _infer_field_type
- api_create_table_from_xlsx
"""
import json
import logging
import re
import secrets
import string
from datetime import datetime, timezone as dt_timezone

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from accounts.rate_limit import rate_limit

from idcards.models import IDCardTable, IDCard
from ..services import IDCardService
from ..services.permission_service import api_require_permission, PermissionService

from .idcard_helpers import (
    _safe_error,
    _check_client_scope_by_group,
    _check_client_scope_by_table,
)

# Logger for this module
logger = logging.getLogger(__name__)
_TABLE_DELETE_CODE_TTL_SECONDS = 600
_TABLE_DELETE_MAX_ATTEMPTS = 5


def _table_delete_code_session_key(table_id: int) -> str:
    return f'table_delete_code_{table_id}'


def _table_delete_attempts_session_key(table_id: int) -> str:
    return f'table_delete_attempts_{table_id}'


def _read_table_delete_attempts(request, table_id: int) -> int:
    try:
        return int(request.session.get(_table_delete_attempts_session_key(table_id), 0) or 0)
    except (TypeError, ValueError):
        return 0


def _set_table_delete_attempts(request, table_id: int, attempts: int) -> None:
    request.session[_table_delete_attempts_session_key(table_id)] = max(int(attempts or 0), 0)
    request.session.modified = True


def _reset_table_delete_attempts(request, table_id: int) -> None:
    key = _table_delete_attempts_session_key(table_id)
    if key in request.session:
        del request.session[key]
        request.session.modified = True


def _increment_table_delete_attempts(request, table_id: int) -> int:
    attempts = _read_table_delete_attempts(request, table_id) + 1
    _set_table_delete_attempts(request, table_id, attempts)
    return attempts


def _store_table_delete_code(request, table_id: int, code: str) -> None:
    request.session[_table_delete_code_session_key(table_id)] = {
        'code': str(code or ''),
        'generated_at': datetime.now(dt_timezone.utc).isoformat(),
    }
    _reset_table_delete_attempts(request, table_id)
    request.session.modified = True


def _consume_table_delete_code_if_valid(request, table_id: int, provided_code: str):
    payload = request.session.get(_table_delete_code_session_key(table_id))
    now = datetime.now(dt_timezone.utc)

    if not isinstance(payload, dict):
        _increment_table_delete_attempts(request, table_id)
        return False, 'missing'

    generated_raw = str(payload.get('generated_at') or '').strip()
    generated_at = None
    if generated_raw:
        try:
            generated_at = datetime.fromisoformat(generated_raw.replace('Z', '+00:00'))
            if generated_at.tzinfo is None:
                generated_at = generated_at.replace(tzinfo=dt_timezone.utc)
        except Exception:
            generated_at = None

    if not generated_at or (now - generated_at).total_seconds() > _TABLE_DELETE_CODE_TTL_SECONDS:
        request.session.pop(_table_delete_code_session_key(table_id), None)
        request.session.modified = True
        _increment_table_delete_attempts(request, table_id)
        return False, 'expired'

    expected = str(payload.get('code') or '')
    if expected != str(provided_code or ''):
        _increment_table_delete_attempts(request, table_id)
        return False, 'invalid'

    request.session.pop(_table_delete_code_session_key(table_id), None)
    _reset_table_delete_attempts(request, table_id)
    request.session.modified = True
    return True, 'ok'


def _extract_confirmation_code(request) -> str:
    if request.content_type and 'application/json' in request.content_type:
        try:
            body = json.loads(request.body or '{}')
            return str((body or {}).get('confirmation_code', '') or '').strip()
        except (json.JSONDecodeError, TypeError, ValueError):
            return ''
    return str(request.POST.get('confirmation_code', '') or '').strip()


# ==================== ID CARD TABLE API ENDPOINTS ====================

@require_http_methods(["POST"])
@api_require_permission('perm_idcard_setting_add')
def api_idcard_table_create(request, group_id):
    """API endpoint to create a new ID Card Table"""
    group, err = _check_client_scope_by_group(request.user, group_id)
    if err: return err
    try:
        data = json.loads(request.body)
        result = IDCardService.create_table(group_id, data)
        return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data!'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["GET"])
@api_require_permission('perm_idcard_setting_list')
@rate_limit(max_requests=60, window_seconds=60, key_prefix='idcard_table_get')
def api_idcard_table_get(request, table_id):
    """API endpoint to get a single ID Card Table"""
    table, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    result = IDCardService.get_table(table_id)
    return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)


@require_http_methods(["POST", "PUT"])
@api_require_permission('perm_idcard_setting_edit')
def api_idcard_table_update(request, table_id):
    """API endpoint to update an ID Card Table"""
    table, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    try:
        data = json.loads(request.body)
        result = IDCardService.update_table(table_id, data)
        return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data!'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["DELETE", "POST"])
@api_require_permission('perm_idcard_setting_delete')
def api_idcard_table_delete(request, table_id):
    """API endpoint to delete an ID Card Table.

    Client / client_staff users: *soft-delete* — sets deleted_by_client=True.
    The table becomes invisible to them but is still fully visible to admin
    as "User Deleted".

    Admin / admin_staff users: *hard-delete* — permanently removes the table
    and all associated card data.
    """
    table, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    try:
        # Prevent deleting tables that still contain cards. Require explicit
        # cleanup (move/delete) of cards before allowing table deletion.
        if IDCard.objects.filter(table_id=table_id).exists():
            return JsonResponse({
                'success': False,
                'message': 'Cannot delete table: it contains cards. Please move or delete all cards first.'
            }, status=400)
        # Client users: soft-delete only
        if PermissionService.is_client_role(request.user):
            table.deleted_by_client = True
            table.save(update_fields=['deleted_by_client'])
            return JsonResponse({'success': True, 'message': 'Table removed from your view successfully.'})

        # Admin-staff hard-delete requires one-time 10-digit confirmation code.
        if PermissionService.is_admin_staff(request.user):
            attempts = _read_table_delete_attempts(request, table_id)
            if attempts >= _TABLE_DELETE_MAX_ATTEMPTS:
                return JsonResponse({
                    'success': False,
                    'message': 'Too many failed attempts. Generate a fresh 10-digit confirmation code.',
                    'code': 'too_many_attempts',
                }, status=429)

            confirmation_code = _extract_confirmation_code(request)
            if len(confirmation_code) != 10 or not confirmation_code.isdigit():
                return JsonResponse({
                    'success': False,
                    'message': '10-digit confirmation code is required for table deletion.',
                    'code': 'confirmation_required',
                }, status=400)

            is_valid, reason = _consume_table_delete_code_if_valid(request, table_id, confirmation_code)
            if not is_valid:
                message = 'Invalid confirmation code. Generate a fresh code and retry.'
                code = 'invalid_confirmation_code'
                if reason == 'expired':
                    message = 'Confirmation code expired. Generate a fresh code and retry.'
                    code = 'expired_confirmation_code'
                return JsonResponse({'success': False, 'message': message, 'code': code}, status=400)

        # Admin / admin_staff: hard-delete
        result = IDCardService.delete_table(table_id)
        return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)
    except Exception as e:
        logger.exception("Table delete error: %s", e)
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["POST"])
@api_require_permission('perm_idcard_setting_delete')
def api_generate_table_delete_code(request, table_id):
    """Generate a 10-digit confirmation code for hard table delete."""
    table, err = _check_client_scope_by_table(request.user, table_id)
    if err:
        return err

    # Client roles only soft-delete and do not need hard-delete code generation.
    if PermissionService.is_client_role(request.user):
        return JsonResponse({'success': False, 'message': 'Confirmation code is not required for client table removal.'}, status=400)

    code = ''.join(secrets.choice(string.digits) for _ in range(10))
    _store_table_delete_code(request, table_id, code)
    return JsonResponse({
        'success': True,
        'code': code,
        'table_id': table.id,
        'table_name': table.name,
    })


@require_http_methods(["POST"])
@api_require_permission('perm_idcard_setting_status')
def api_idcard_table_toggle_status(request, table_id):
    """API endpoint to toggle ID Card Table active/inactive status"""
    table, err = _check_client_scope_by_table(request.user, table_id)
    if err: return err
    try:
        result = IDCardService.toggle_table_status(table_id)
        return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)
    except Exception as e:
        logger.exception("Table toggle status error: %s", e)
        return JsonResponse({'success': False, 'message': _safe_error(e)}, status=500)


@require_http_methods(["GET"])
@api_require_permission('perm_idcard_setting_list')
def api_idcard_table_list(request, group_id):
    """API endpoint to list all ID Card Tables for a group"""
    group, err = _check_client_scope_by_group(request.user, group_id)
    if err: return err
    result = IDCardService.list_tables(group_id)
    return JsonResponse(result.to_response_dict(), status=200 if result.success else 400)


# ==================== CREATE TABLE FROM XLSX ====================

# Field-type inference patterns: map XLSX header names to field types.
# Order matters: first match wins.  All comparisons are case-insensitive.
_HEADER_TYPE_MAP = [
    # Image-like fields
    (['mother photo', 'm photo', 'mother_photo', 'mother pic'], 'rel_photo'),
    (['father photo', 'f photo', 'father_photo', 'father pic'], 'rel_photo'),
    (['relation photo', 'relation image', 'relation pic', 'rel photo'], 'rel_photo'),
    (['relation 1 photo', 'relation1photo', 'relation one photo', 'rel 1 photo', 'rel1photo', 'rel_1photo'], 'rel_photo'),
    (['relation 2 photo', 'relation2photo', 'relation two photo', 'rel 2 photo', 'rel2photo', 'rel_2photo'], 'rel_photo'),
    (['photo', 'pic', 'picture', 'image', 'student photo', 'student image'], 'photo'),
    (['signature', 'sign'], 'signature'),
    (['barcode'], 'barcode'),
    (['qr code', 'qr_code', 'qr'], 'qr_code'),
    # Structural fields
    (['class'], 'class'),
    (['section', 'sec'], 'section'),
    (['email', 'e-mail', 'email id', 'email address'], 'email'),
]



_SAMPLE_ROW_SCAN_LIMIT = 3
_RELATION_SLOT_HEADER_RE = re.compile(
    r'^(?:rel(?:ation)?)\s*[_-]?\s*(?:1|one|2|two)$'
)
_REL_PHOTO_HEADER_RE = re.compile(
    r'^(?:rel(?:ation)?)\s*[_-]?\s*(?:1|one|2|two)\s*(?:photo|image|pic|picture)$'
)
_IMAGE_EXTENSION_RE = re.compile(r'\.(?:jpe?g|png|gif|bmp|webp|heic|heif)$', re.IGNORECASE)


def _is_relation_slot_header(normalized_header: str) -> bool:
    """Return True for plain relation slots like REL_1 / REL_2 (without photo keyword)."""
    if not normalized_header:
        return False
    return bool(_RELATION_SLOT_HEADER_RE.match(normalized_header))


def _is_relation_photo_header(normalized_header: str) -> bool:
    """Return True when XLSX header clearly indicates relation photo columns."""
    if not normalized_header:
        return False
    if _REL_PHOTO_HEADER_RE.match(normalized_header):
        return True
    return bool(re.search(r'\b(?:father|mother)\b\s*(?:photo|image|pic|picture)', normalized_header))


def _is_image_like_relation_value(value) -> bool:
    """Heuristic: image identifiers are usually ids/paths/filenames, not person names."""
    if value is None:
        return False

    text = str(value).strip()
    if not text:
        return False

    normalized = text.lower().strip().replace('\\', '/')
    if normalized.startswith('pending:'):
        normalized = normalized[8:].strip()

    if _IMAGE_EXTENSION_RE.search(normalized):
        return True

    if normalized.startswith(('http://', 'https://')) and (
        '/media/' in normalized or '/card_media/' in normalized or _IMAGE_EXTENSION_RE.search(normalized)
    ):
        return True

    if '/' in normalized and re.search(r'(?:photo|image|pic|card_media|id_photos)', normalized):
        return True

    compact = re.sub(r'[\s_-]+', '', normalized)
    if re.match(r'^(?:img|image|photo|pic)\d{3,}$', compact):
        return True
    if re.match(r'^[a-z]?\d{6,}$', compact):
        return True

    return False


def _is_text_like_relation_value(value) -> bool:
    """Heuristic: relation-name fields usually contain alphabetic words."""
    if value is None:
        return False

    text = str(value).strip()
    if not text:
        return False
    if _is_image_like_relation_value(text):
        return False

    letters = len(re.findall(r'[A-Za-z]', text))
    digits = len(re.findall(r'\d', text))

    if letters >= 2 and (digits == 0 or letters >= digits):
        return True
    if ' ' in text and letters >= 1:
        return True
    return False


def _infer_relation_slot_type(sample_values) -> str:
    """Infer REL_1/REL_2 slot type using first few data rows."""
    image_votes = 0
    text_votes = 0

    for value in (sample_values or [])[:_SAMPLE_ROW_SCAN_LIMIT]:
        if _is_image_like_relation_value(value):
            image_votes += 1
            continue
        if _is_text_like_relation_value(value):
            text_votes += 1

    if image_votes > 0 and image_votes >= text_votes:
        return 'rel_photo'
    return 'text'


def _infer_field_type(header_name: str, sample_values=None) -> str:
    """Infer field type from an XLSX header name.

    Returns one of the VALID_FIELD_TYPES for IDCardTable.
    Falls back to 'text' for any unrecognised header.
    """
    normalized = header_name.strip().lower().replace('_', ' ')
    normalized = re.sub(r'\s+', ' ', normalized).strip()
    if _is_relation_photo_header(normalized):
        return 'rel_photo'
    if _is_relation_slot_header(normalized):
        return _infer_relation_slot_type(sample_values)
    for patterns, field_type in _HEADER_TYPE_MAP:
        if normalized in patterns:
            return field_type
    return 'text'


@require_http_methods(["POST"])
@api_require_permission('perm_idcard_setting_add')
def api_create_table_from_xlsx(request, group_id):
    """
    Create a new IDCardTable from an XLSX file's header row, then bulk-upload
    the data rows into the table.  Optionally accepts ZIP files for image fields.

    This combines two steps into one:
      1. Reads the first row of the XLSX to derive field names + types.
      2. Creates a table with those fields.
      3. Delegates to the existing bulk-upload logic to import the data.

    POST /api/group/<group_id>/table/create-from-xlsx/

    Form-data:
        file              : XLSX/XLS/CSV file   (required)
        table_name        : Optional override for the table name
        photos_zip_<FIELD>: ZIP per image field  (optional)
        unified_zip_<N>   : Unified ZIPs         (optional)
        unified_zip_count : Number of unified ZIPs (optional)
        zip_field_names   : JSON array of field names with ZIP uploads (optional)

    Returns JSON:
        { success, message, table_id, table_name, cards_created, ... }
    """
    from io import BytesIO

    # Admin-only hard gate: clients/client_staff should never use this flow.
    if PermissionService.is_client_role(request.user):
        return JsonResponse({
            'success': False,
            'message': 'Create with XLSX is not available for client accounts.'
        }, status=403)

    group, err = _check_client_scope_by_group(request.user, group_id)
    if err:
        return err

    # ── 1. Validate file ────────────────────────────────────────────
    if 'file' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No file uploaded.'}, status=400)

    uploaded_file = request.FILES['file']
    file_name = uploaded_file.name.lower()
    if not file_name.endswith(('.xlsx', '.xls', '.csv')):
        return JsonResponse({
            'success': False,
            'level': 'warning',
            'message': 'Only .xlsx, .xls, and .csv files are supported.'
        }, status=400)

    # Size guard: 50 MB max for the spreadsheet
    if uploaded_file.size > 50 * 1024 * 1024:
        return JsonResponse({
            'success': False,
            'level': 'warning',
            'message': 'Spreadsheet file must be under 50 MB.'
        }, status=400)

    # ── 2. Read headers ─────────────────────────────────────────────
    sample_rows = []

    def _clean_spreadsheet_cell(cell):
        if isinstance(cell, str):
            return (
                cell.strip()
                .replace('_x000D_', '')
                .replace('_X000D_', '')
                .replace('_x000d_', '')
                .replace('\r', '')
            )
        return cell

    try:
        if file_name.endswith('.csv'):
            import csv, io
            content = uploaded_file.read().decode('utf-8-sig', errors='replace')
            uploaded_file.seek(0)
            reader = csv.reader(io.StringIO(content))
            headers = next(reader, [])
            for row in reader:
                if len(sample_rows) >= _SAMPLE_ROW_SCAN_LIMIT:
                    break
                sample_rows.append(list(row))
        else:
            file_content = uploaded_file.read()
            uploaded_file.seek(0)

            if len(file_content) < 4:
                raise ValueError('Spreadsheet file is too small.')

            magic_bytes = file_content[:4]
            is_zip = magic_bytes[:2] == b'PK'
            is_old_xls = (magic_bytes[0] == 0xD0 and magic_bytes[1] == 0xCF)

            if is_old_xls or file_name.endswith('.xls'):
                import xlrd

                wb = xlrd.open_workbook(file_contents=file_content)
                ws = wb.sheet_by_index(0)

                headers = [
                    _clean_spreadsheet_cell(ws.cell_value(0, col_idx)) if ws.nrows > 0 else ''
                    for col_idx in range(ws.ncols)
                ]

                max_row = min(ws.nrows, 1 + _SAMPLE_ROW_SCAN_LIMIT)
                for row_idx in range(1, max_row):
                    row = [
                        _clean_spreadsheet_cell(ws.cell_value(row_idx, col_idx))
                        for col_idx in range(ws.ncols)
                    ]
                    sample_rows.append(row)
            elif is_zip or file_name.endswith('.xlsx'):
                import openpyxl

                wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=True)
                ws = wb.active
                raw_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                headers = [
                    _clean_spreadsheet_cell(cell) if cell is not None else ''
                    for cell in raw_row
                ]
                for row in ws.iter_rows(min_row=2, max_row=1 + _SAMPLE_ROW_SCAN_LIMIT, values_only=True):
                    sample_rows.append([_clean_spreadsheet_cell(cell) for cell in row])
                wb.close()
            else:
                raise ValueError('Unrecognized spreadsheet format.')
    except Exception as exc:
        logger.error("Failed to read spreadsheet headers: %s", exc)
        return JsonResponse({
            'success': False,
            'message': 'Could not read the spreadsheet. Please check the file format.'
        }, status=400)

    # Filter out empty headers while preserving original column indices.
    indexed_headers = [(idx, h) for idx, h in enumerate(headers) if h]
    headers = [h for _, h in indexed_headers]
    if not headers:
        return JsonResponse({
            'success': False, 'message': 'The spreadsheet has no column headers in the first row.'
        }, status=400)

    if len(headers) > IDCardService.MAX_FIELDS_PER_TABLE:
        return JsonResponse({
            'success': False,
            'level': 'warning',
            'message': f'Maximum {IDCardService.MAX_FIELDS_PER_TABLE} columns allowed. '
                       f'Your file has {len(headers)} columns.'
        }, status=400)

    # ── 3. Infer field definitions (or use client-provided config) ──
    field_config_json = request.POST.get('field_config', '')
    client_field_config = None
    if field_config_json:
        try:
            import json as _json
            client_field_config = _json.loads(field_config_json)
            if not isinstance(client_field_config, list):
                client_field_config = None
        except (ValueError, TypeError):
            client_field_config = None

    VALID_FIELD_TYPES = {
        'text', 'class', 'section', 'email', 'photo',
        'rel_photo',
        # Legacy aliases accepted and normalized to rel_photo.
        'mother_photo', 'father_photo',
        'signature', 'barcode', 'qr_code',
    }

    fields = []
    for idx, header in enumerate(headers):
        source_col_idx = indexed_headers[idx][0]
        sample_values = [
            row[source_col_idx] if source_col_idx < len(row) else None
            for row in sample_rows
        ]
        # Default: auto-infer
        field_type = _infer_field_type(header, sample_values=sample_values)
        mandatory = False
        # Default field name from the XLSX header
        field_name = header.strip().upper()

        # Override with client-provided config if available and valid
        if client_field_config and idx < len(client_field_config):
            cfg = client_field_config[idx]
            if isinstance(cfg, dict):
                cfg_type = cfg.get('type', '')
                if cfg_type in VALID_FIELD_TYPES:
                    field_type = 'rel_photo' if cfg_type in ('mother_photo', 'father_photo') else cfg_type
                mandatory = bool(cfg.get('mandatory', False))
                # Accept client-provided field name (user may have edited it in preview)
                cfg_name = cfg.get('name', '')
                if isinstance(cfg_name, str):
                    cfg_name = cfg_name.strip().upper()
                    if cfg_name:
                        field_name = cfg_name

        fields.append({
            'name': field_name,
            'type': 'rel_photo' if field_type in ('mother_photo', 'father_photo') else field_type,
            'order': idx,
            'mandatory': mandatory,
        })

    # ── 4. Create the table ─────────────────────────────────────────
    table_name = (request.POST.get('table_name') or '').strip().upper()
    if not table_name:
        # Derive from filename (strip extension)
        import os as _os
        base = _os.path.splitext(uploaded_file.name)[0]
        table_name = base.strip().upper()[:255] or 'IMPORTED TABLE'

    try:
        table = IDCardTable.objects.create(
            group=group,
            name=table_name,
            fields=fields,
            is_active=True,
        )
    except Exception as exc:
        logger.exception("Failed to create table from XLSX: %s", exc)
        return JsonResponse({
            'success': False, 'message': 'Failed to create table. Please try again.'
        }, status=500)

    logger.info(
        "Created table %d (%s) with %d fields from XLSX (user=%s)",
        table.id, table_name, len(fields), request.user.id,
    )

    # ── 5. Delegate to existing bulk upload ─────────────────────────
    # We re-use the synchronous bulk upload by faking the table_id into the
    # existing function.  The uploaded file + ZIPs are already in request.FILES.
    try:
        # Import and call the existing bulk upload handler directly, passing our
        # newly-created table_id.  We intercept its JsonResponse to enrich it.
        from core.views.idcard_api import api_idcard_bulk_upload as _bulk_upload

        # Temporarily patch the URL kwargs so the bulk upload sees our table
        bulk_response = _bulk_upload(request, table.id)

        # Parse the JSON body from the upload response
        import json as _json
        try:
            body = _json.loads(bulk_response.content)
        except Exception:
            body = {}

        if bulk_response.status_code == 200 and body.get('success'):
            body['table_id'] = table.id
            body['table_name'] = table.name
            body['fields_created'] = len(fields)
            body['message'] = (
                f'Table "{table.name}" created with {len(fields)} fields. '
                + (body.get('message') or
                   f'{body.get("cards_created", 0)} cards imported.')
            )
            return JsonResponse(body)
        else:
            # Upload failed — clean up: delete the empty table
            try:
                table.delete()
            except Exception:
                pass
            return bulk_response

    except Exception as exc:
        logger.exception("Bulk upload after table creation failed: %s", exc)
        # Clean up table
        try:
            table.delete()
        except Exception:
            pass
        return JsonResponse({
            'success': False,
            'message': 'Table was created but data import failed. Please try again.'
        }, status=500)
