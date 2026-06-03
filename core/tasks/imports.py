import io
import os
import zipfile
import openpyxl
from celery import shared_task
from django.db import transaction
from django.utils.text import slugify
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from core.models import Tenant, UserProfile, DynamicTable, TableField, CardRecord, Job, JobLog
from core.services.storage import StorageService
from core.services.image import ImageService
from core.services.table import TableService
from core.validators import DynamicSchemaValidator

# =====================================================================
# helper functions for job progress logging
# =====================================================================

def log_job(job: Job, message: str, level: str = 'INFO'):
    JobLog.objects.create(job=job, message=message, level=level)

def update_progress(job: Job, progress: int):
    job.progress = progress
    job.save(update_fields=['progress', 'updated_at'])

# =====================================================================
# IMPORT TYPE 1: Create Table From Excel Schema
# =====================================================================

@shared_task(bind=True)
def import_xlsx_schema_task(self, job_id: str):
    """
    Parses the Excel file, reads headers, creates a DynamicTable and fields, 
    and inserts rows as PENDING card records.
    """
    job = Job.objects.get(id=job_id)
    job.status = 'RUNNING'
    job.save()
    
    storage = StorageService()
    
    try:
        # Load parameters from job payload
        xlsx_path = job.payload.get('xlsx_path')
        table_name = job.payload.get('table_name')
        client_profile_id = job.payload.get('client_profile_id')
        
        client_profile = UserProfile.objects.get(id=client_profile_id)
        
        log_job(job, "Reading Excel schema data...")
        xlsx_file = storage.read_file(xlsx_path)
        wb = openpyxl.load_workbook(xlsx_file, data_only=True)
        sheet = wb.active
        
        # Phase 1: Determine dynamic schema from headers (Row 1)
        headers = [cell.value for cell in sheet[1] if cell.value is not None]
        fields_schema = []
        for header in headers:
            key = slugify(header).replace('-', '_')
            # Default to TEXT type for raw excel imports
            fields_schema.append({
                'name': header,
                'key': key,
                'type': 'TEXT',
                'is_required': False
            })
            
        log_job(job, f"Creating dynamic table '{table_name}' with {len(fields_schema)} fields...")
        table = TableService.create_table(job.tenant, client_profile, table_name, fields_schema)
        
        # Phase 2: Populate rows
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        total_rows = len(rows)
        log_job(job, f"Importing {total_rows} records...")
        
        for idx, row in enumerate(rows):
            record_data = {}
            for col_idx, cell_value in enumerate(row):
                if col_idx < len(fields_schema):
                    key = fields_schema[col_idx]['key']
                    record_data[key] = str(cell_value) if cell_value is not None else ""
                    
            CardRecord.objects.create(
                table=table,
                data=record_data,
                status='PENDING'
            )
            
            # Progress update
            progress_pct = int(((idx + 1) / total_rows) * 100)
            if progress_pct % 10 == 0:
                update_progress(job, progress_pct)
                
        job.status = 'COMPLETED'
        job.save()
        log_job(job, "Schema import completed successfully.")
        storage.delete_file(xlsx_path)
        
    except Exception as e:
        job.status = 'FAILED'
        job.error_message = str(e)
        job.save()
        log_job(job, f"Fatal error during schema creation: {str(e)}", 'ERROR')


# =====================================================================
# IMPORT TYPE 2: XLSX + ZIP Images (Multi-Phase)
# =====================================================================

@shared_task(bind=True)
def import_xlsx_zip_task(self, job_id: str):
    """
    Import Type 2: Multi-phase XLSX + ZIP matched card record creations.
    """
    job = Job.objects.get(id=job_id)
    job.status = 'RUNNING'
    job.save()
    
    storage = StorageService()
    img_service = ImageService(storage)
    
    try:
        # Load parameters
        xlsx_path = job.payload.get('xlsx_path')
        zip_path = job.payload.get('zip_path')
        table_id = job.payload.get('table_id')
        
        table = DynamicTable.objects.get(id=table_id)
        fields = {f.name: f for f in table.fields.all()}
        
        # -------------------------------------------------------------
        # PHASE 1: Validate XLSX (0% - 15%)
        # -------------------------------------------------------------
        log_job(job, "Phase 1: Validating XLSX spreadsheet structure...")
        update_progress(job, 5)
        xlsx_file = storage.read_file(xlsx_path)
        wb = openpyxl.load_workbook(xlsx_file, data_only=True)
        sheet = wb.active
        
        headers = [cell.value for cell in sheet[1]]
        # Confirm required headers exist
        for f_name, field in fields.items():
            if field.is_required and f_name not in headers:
                raise ValidationError(f"Missing required spreadsheet column: '{f_name}'")
                
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        total_rows = len(rows)
        log_job(job, f"Validated XLSX. Total import row count: {total_rows}")
        update_progress(job, 15)
        
        # -------------------------------------------------------------
        # PHASE 2: Validate ZIP (15% - 30%)
        # -------------------------------------------------------------
        log_job(job, "Phase 2: Inspecting ZIP archive layout...")
        zip_bytes = storage.read_file(zip_path).read()
        archive = zipfile.ZipFile(io.BytesIO(zip_bytes))
        zip_files = set(archive.namelist())
        log_job(job, f"ZIP contents verified. Found {len(zip_files)} files.")
        update_progress(job, 30)
        
        # -------------------------------------------------------------
        # PHASE 3: Match Images (30% - 45%)
        # -------------------------------------------------------------
        log_job(job, "Phase 3: Cross-matching Excel values with ZIP filenames...")
        matched_images = {}  # RowIndex -> ColumnKey -> ZipFileName
        missing_images_count = 0
        
        for idx, row in enumerate(rows):
            for col_idx, cell_value in enumerate(row):
                if col_idx >= len(headers):
                    continue
                header = headers[col_idx]
                field = fields.get(header)
                
                if field and field.type == 'IMAGE' and cell_value:
                    image_filename = str(cell_value).strip()
                    if image_filename in zip_files:
                        matched_images.setdefault(idx, {})[field.key] = image_filename
                    else:
                        missing_images_count += 1
                        log_job(job, f"Row {idx+2}: Defined image '{image_filename}' not found in ZIP.", 'WARNING')
                        
        log_job(job, f"Image matching complete. Missing images: {missing_images_count}.")
        update_progress(job, 45)
        
        # -------------------------------------------------------------
        # PHASE 4: Upload Images (45% - 80%)
        # -------------------------------------------------------------
        log_job(job, "Phase 4: Resizing and uploading matched images...")
        uploaded_metadata = {}  # RowIndex -> ColumnKey -> ImageMetadataDict
        
        for idx in range(total_rows):
            if idx in matched_images:
                for f_key, filename in matched_images[idx].items():
                    with archive.open(filename) as file_in_zip:
                        file_data = io.BytesIO(file_in_zip.read())
                        # Run compression, scaling, and upload drivers
                        metadata = img_service.process_and_upload(file_data, str(table.id), f_key)
                        uploaded_metadata.setdefault(idx, {})[f_key] = metadata
            
            # Progress scale: 45% -> 80%
            current_progress = int(45 + ((idx + 1) / total_rows) * 35)
            if current_progress % 5 == 0:
                update_progress(job, current_progress)
                
        log_job(job, "All matched images resized and uploaded.")
        update_progress(job, 80)
        
        # -------------------------------------------------------------
        # PHASE 5: Create Cards (80% - 100%)
        # -------------------------------------------------------------
        log_job(job, "Phase 5: Writing card records to Database...")
        
        with transaction.atomic():
            for idx, row in enumerate(rows):
                record_data = {}
                for col_idx, cell_value in enumerate(row):
                    if col_idx >= len(headers):
                        continue
                    header = headers[col_idx]
                    field = fields.get(header)
                    if field and field.type != 'IMAGE':
                        record_data[field.key] = cell_value
                        
                images_payload = uploaded_metadata.get(idx, {})
                CardRecord.objects.create(
                    table=table,
                    data=record_data,
                    images=images_payload,
                    status='PENDING'
                )
                
                # Progress scale: 80% -> 100%
                current_progress = int(80 + ((idx + 1) / total_rows) * 20)
                if current_progress % 5 == 0:
                    update_progress(job, current_progress)
                    
        job.status = 'COMPLETED'
        job.save()
        log_job(job, "Pipeline run finished successfully.")
        
        # Cleanup files
        storage.delete_file(xlsx_path)
        storage.delete_file(zip_path)
        
    except Exception as e:
        job.status = 'FAILED'
        job.error_message = str(e)
        job.save()
        log_job(job, f"Fatal error during XLSX+ZIP pipeline execution: {str(e)}", 'ERROR')


# =====================================================================
# IMPORT TYPE 3: Image Bulk Re-upload
# =====================================================================

@shared_task(bind=True)
def reupload_images_zip_task(self, job_id: str):
    """
    Import Type 3: Bulk overwrite image properties by matching zip contents
    against card records via a lookup key (e.g. employee_id = filename).
    """
    job = Job.objects.get(id=job_id)
    job.status = 'RUNNING'
    job.save()
    
    storage = StorageService()
    img_service = ImageService(storage)
    
    try:
        zip_path = job.payload.get('zip_path')
        table_id = job.payload.get('table_id')
        lookup_field_key = job.payload.get('lookup_field_key')
        target_image_field_key = job.payload.get('target_image_field_key')
        
        table = DynamicTable.objects.get(id=table_id)
        
        log_job(job, "Opening target ZIP files...")
        zip_bytes = storage.read_file(zip_path).read()
        archive = zipfile.ZipFile(io.BytesIO(zip_bytes))
        zip_files = archive.namelist()
        total_files = len(zip_files)
        
        log_job(job, f"Matching {total_files} images to existing records...")
        
        for idx, filename in enumerate(zip_files):
            # Parse prefix/filename to determine lookup ID key (e.g. "EMP1024.jpg" -> "EMP1024")
            lookup_val, _ = os.path.splitext(os.path.basename(filename))
            
            # Query card record inside JSONB data column
            record = CardRecord.objects.filter(
                table=table,
                data__contains={lookup_field_key: lookup_val}
            ).first()
            
            if record:
                # Remove existing images if present
                old_metadata = record.images.get(target_image_field_key)
                if old_metadata:
                    img_service.delete_images(old_metadata)
                    
                # Upload replacement image
                with archive.open(filename) as zip_img_file:
                    img_io = io.BytesIO(zip_img_file.read())
                    new_metadata = img_service.process_and_upload(img_io, str(table.id), target_image_field_key)
                    
                record.images[target_image_field_key] = new_metadata
                record.version += 1
                record.save()
            else:
                log_job(job, f"No matching record found for value: '{lookup_val}' (file: {filename})", 'WARNING')
                
            progress_pct = int(((idx + 1) / total_files) * 100)
            if progress_pct % 10 == 0:
                update_progress(job, progress_pct)
                
        job.status = 'COMPLETED'
        job.save()
        log_job(job, "Image re-upload process finished.")
        storage.delete_file(zip_path)
        
    except Exception as e:
        job.status = 'FAILED'
        job.error_message = str(e)
        job.save()
        log_job(job, f"Fatal error during image re-upload: {str(e)}", 'ERROR')
